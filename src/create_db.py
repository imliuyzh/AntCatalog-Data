import sqlite3
import openpyxl
from pathlib import Path
from time import asctime

class DBCreator:
    '''A class to facilitate the database creation process.'''

    def __init__(self):
        '''Constructor for the DBCreator class. self._count is used to set the ID for each class.'''
        self._count = 0

    def run(self) -> None:
        '''Create the schemas, import the data from spreadsheets, create a view for every instructor in a course, and finalize the indexes.'''
        self._create_table()
        self._insert_data()
        self._create_view()
        self._create_index()

    def _create_table(self) -> None:
        '''Create a file named data.db and create two SQL tables.'''
        connection = sqlite3.connect("../data.db")
        try:
            print(f"[{asctime()}] Begin creating table")
            connection.executescript(
                """
                CREATE TABLE Course (
                    course_id INTEGER CHECK (typeof(course_id) = 'integer'),
                    year INTEGER NOT NULL CHECK (typeof(year) = 'integer'),
                    quarter TEXT NOT NULL CHECK (typeof(quarter) = 'text'),
                    course_code INTEGER NOT NULL CHECK (typeof(course_code) = 'integer'),
                    department TEXT NOT NULL CHECK (typeof(department) = 'text'),
                    course_number TEXT NOT NULL CHECK (typeof(course_number) = 'text'),
                    course_title TEXT NOT NULL CHECK (typeof(course_title) = 'text'),
                    grade_a_count INTEGER NOT NULL CHECK (typeof(grade_a_count) = 'integer'),
                    grade_b_count INTEGER NOT NULL CHECK (typeof(grade_b_count) = 'integer'),
                    grade_c_count INTEGER NOT NULL CHECK (typeof(grade_c_count) = 'integer'),
                    grade_d_count INTEGER NOT NULL CHECK (typeof(grade_d_count) = 'integer'),
                    grade_f_count INTEGER NOT NULL CHECK (typeof(grade_f_count) = 'integer'),
                    grade_p_count INTEGER NOT NULL CHECK (typeof(grade_p_count) = 'integer'),
                    grade_np_count INTEGER NOT NULL CHECK (typeof(grade_np_count) = 'integer'),
                    gpa_avg REAL NOT NULL CHECK (typeof(gpa_avg) = 'real'),
                    CONSTRAINT CoursePrimaryKey PRIMARY KEY (course_id)
                );

                CREATE TABLE Instructor (
                    course_id INTEGER CHECK (typeof(course_id) = 'integer'),
                    name TEXT CHECK (typeof(name) = 'text'),
                    CONSTRAINT InstructorPrimaryKey PRIMARY KEY (course_id, name),
                    CONSTRAINT InstructorCourseIdForeignKey FOREIGN KEY (course_id) REFERENCES Course(course_id) ON DELETE CASCADE
                );
                """
            )
            print(f"[{asctime()}] Finished creating table")
            connection.commit()
        except sqlite3.Error as error:
            print(f"[{asctime()}] Failed to create table: {error}")
            connection.rollback()
        finally:
            if connection is not None:
                connection.close()

    def _load_data_per_year(self, workbook: openpyxl.workbook.workbook.Workbook) -> ([(int, str, int, str, str, str, int, int, int, int, int, int, int, float)], [(str, int, str)]):
        '''Return data in an academic year.'''
        all_courses, all_instructors = [], []
        for term in workbook.sheetnames:
            print(f"[{asctime()}] Processing data from {term}")
            sheet = workbook[term]
            
            for row in range(2, sheet.max_row + 1):
                quarter, year = str(term.split()[0]), int(term.split()[1])
                department = str(sheet["C" + str(row)].value)
                course_number = str(sheet["D" + str(row)].value)
                course_code = int(sheet["E" + str(row)].value)
                course_title = str(sheet["F" + str(row)].value)
                grade_a_count = int(sheet["H" + str(row)].value)
                grade_b_count = int(sheet["I" + str(row)].value)
                grade_c_count = int(sheet["J" + str(row)].value)
                grade_d_count = int(sheet["K" + str(row)].value)
                grade_f_count = int(sheet["L" + str(row)].value)
                grade_p_count = int(sheet["M" + str(row)].value)
                grade_np_count = int(sheet["N" + str(row)].value)
                average_gpa = float(sheet["O" + str(row)].value) if sheet["O" + str(row)].value is not None else 0
                instructors = sheet["G" + str(row)].value.split("; ")
                
                all_courses.append((self._count, year, quarter, course_code, department.upper(), course_number.upper(), course_title.upper(), grade_a_count, grade_b_count, grade_c_count, grade_d_count, grade_f_count, grade_p_count, grade_np_count, average_gpa))
                for instructor in instructors:
                    all_instructors.append((self._count, instructor.upper()))
                
                self._count += 1

            print(f"[{asctime()}] Finished processing data from {term}")
        return all_courses, all_instructors

    def _insert_data(self) -> None:
        '''Loop through the spreadsheets to parse and store all the course statistics into data.db.'''
        folder = Path(r"../processed_data")
        connection = sqlite3.connect("../data.db")

        try:
            for spreadsheet_path in folder.iterdir():
                print(f"[{asctime()}] Working with the file {spreadsheet_path}")
                courses, instructors = self._load_data_per_year(openpyxl.load_workbook(str(spreadsheet_path), data_only=True))
                connection.executemany("INSERT INTO Course VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);", courses)
                connection.executemany("INSERT INTO Instructor VALUES (?, ?);", instructors)
                connection.commit()
        except sqlite3.Error as error:
            print(f"[{asctime()}] Failed to insert data: {error}")
            connection.rollback()
        finally:
            if connection is not None:
                connection.close()
                
    def _create_view(self) -> None:
        '''Create a view storing a string for every instructor, separated by /, of all courses.'''
        connection = sqlite3.connect("../data.db")
        try:
            print(f"[{asctime()}] Begin creating a view")
            connection.executescript(
                """
                CREATE VIEW InstructorView AS
                    SELECT
                        I.course_id,
                        GROUP_CONCAT(I.name, '/') AS names
                    FROM Instructor I
                    GROUP BY I.course_id;
                """
            )
            print(f"[{asctime()}] Finished creating a view")
            connection.commit()
        except sqlite3.Error as error:
            print(f"[{asctime()}] Failed to create a view: {error}")
            connection.rollback()
        finally:
            if connection is not None:
                connection.close()
    
    def _create_index(self) -> None:
        '''Index several fields on the Course and Instructor table.'''
        connection = sqlite3.connect("../data.db")
        try:
            print(f"[{asctime()}] Begin creating indices")
            connection.executescript(
                """
                CREATE INDEX CourseCodeCourseNumberDepartmentQuarterYearndex ON Course (course_code, course_number, department, quarter, year);
                CREATE INDEX CourseCodeCourseNumberQuarterYearndex ON Course (course_code, course_number, quarter, year);
                CREATE INDEX CourseCodeCourseNumberYearndex ON Course (course_code, course_number, year);
                CREATE INDEX CourseCodeDepartmentQuarterYearndex ON Course (course_code, department, quarter, year);
                CREATE INDEX CourseCodeDepartmentYearndex ON Course (course_code, department, year);
                CREATE INDEX CourseCodeQuarterYearndex ON Course (course_code, quarter, year);
                CREATE INDEX CourseCodeYearndex ON Course (course_code, year);

                CREATE INDEX CourseNumberDepartmentQuarterYearndex ON Course (course_number, department, quarter, year);
                CREATE INDEX CourseNumberDepartmentYearndex ON Course (course_number, department, year);
                CREATE INDEX CourseNumberQuarterYearIndex ON Course (course_number, quarter, year);
                CREATE INDEX CourseNumberYearIndex ON Course (course_number, year);

                CREATE INDEX DepartmentQuarterYearndex ON Course (department, quarter, year);
                CREATE INDEX DepartmentYearIndex ON Course (department, year);

                CREATE INDEX QuarterYearndex ON Course (quarter, year);

                CREATE INDEX YearIndex ON Course (year);
                
                CREATE INDEX InstructorNameIndex ON Instructor (name);
                """
            )
            print(f"[{asctime()}] Finished creating indices")
            connection.commit()
        except sqlite3.Error as error:
            print(f"[{asctime()}] Failed to create index: {error}")
            connection.rollback()
        finally:
            if connection is not None:
                connection.close()

if __name__ == "__main__":
    db_creator = DBCreator()
    db_creator.run()
