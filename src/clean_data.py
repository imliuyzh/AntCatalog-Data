import logging
import openpyxl
import urllib.error, urllib.request
from bs4 import BeautifulSoup
from random import uniform
from time import asctime, sleep

SPREADSHEET_FILES = [] # Change to the name of the file under the "temp" folder you want to parse
TERM_DICT = {
    "FALL": "92",
    "WINTER": "03",
    "SPRING": "14",
    "SUMMER": "25 39 76"
}

def _get_data(requests: [urllib.request.Request], course_code: str) -> dict:
    '''Send requests to https://www.reg.uci.edu/perl/WebSoc?
    Submit=Display+XML+Results&YearTerm=???&Breadth=ANY&Dept=+ALL&Division=ANY&
    ClassType=ALL&FullCourses=ANY&CancelledCourses=Exclude&CourseCodes=???
    in order to parse the XML to get the correct data.'''
    
    info = {
        "success": False,
        "error_message": None,
        "dept_name": None,
        "course_code": course_code,
        "course_number": None,
        "course_title": None,
        "instructors": []
    }

    try:
        for request in requests:
            with urllib.request.urlopen(request) as response:
                content = BeautifulSoup(response.read().decode(), features="xml")
                course_target = content.find("course_code", string=course_code)

                if course_target is not None:
                    if info["success"] == False:
                        info["success"] = True
                        info["dept_name"] = course_target.parent.parent.parent.get("dept_code").strip()
                        info["course_number"] = course_target.parent.parent.get("course_number").strip()
                        info["course_title"] = course_target.parent.parent.get("course_title").strip()

                        prof_list = course_target.find_next_sibling("sec_instructors").contents
                        for prof_element in prof_list:
                            if prof_element != "\n" and prof_element.string.strip() != "STAFF":
                                info["instructors"].append(prof_element.string.strip())
                    else:
                        logging.warning(f"[{asctime()}] There is a collision for Summer course #{course_code} ({sheetname})")
                        break
                    
            sleep(uniform(1, 2))

        if info["success"] == False:
            info["error_message"] = "Failed to find the corresponding course on WebSOC."
    except urllib.error.HTTPError as error_object:
        info["error_message"] = f"Server returns error code ({error_object.code})."
    except urllib.error.URLError as error_object:
        info["error_message"] = error_object.reason
    except TimeoutError as error_object:
        info["error_message"] = error_object

    return info

def _build_summer_requests(course_info: dict) -> [urllib.request.Request]:
    '''Build an list to send correct queries to the WebSOC service for Summer courses.'''
    requests = []
    sessions = TERM_DICT["SUMMER"].split()

    for session in sessions:
        parameters = bytes("Submit=Display+XML+Results&"
                        + f"YearTerm={course_info['year']}-{session}&Breadth=ANY&"
                        + "Dept=+ALL&Division=ANY&ClassType=ALL&"
                        + "FullCourses=ANY&CancelledCourses=Exclude&"
                        + f"CourseCodes={course_info['course_code']}", "utf-8")
        request = urllib.request.Request("https://www.reg.uci.edu/perl/WebSoc", data=parameters)
        request.add_header("User-Agent", "Mozilla/5.0 (Windows NT 6.1; Win64; x64) "
                        + "AppleWebKit/537.36 (KHTML, like Gecko) "
                        + "Chrome/79.0.3945.88 Safari/537.36 Edg/79.0.309.56")
        requests.append(request)
    
    return requests

def _build_non_summer_request(course_info: dict) -> [urllib.request.Request]:
    '''Build an object to send a correct query to the WebSOC service for non-Summer courses.'''
    parameters = bytes("Submit=Display+XML+Results&"
                       + f"YearTerm={course_info['quarter']}&Breadth=ANY&"
                       + "Dept=+ALL&Division=ANY&ClassType=ALL&"
                       + "FullCourses=ANY&CancelledCourses=Exclude&"
                       + f"CourseCodes={course_info['course_code']}", "utf-8")
    request = urllib.request.Request("https://www.reg.uci.edu/perl/WebSoc", data=parameters)
    request.add_header("User-Agent", "Mozilla/5.0 (Windows NT 6.1; Win64; x64) "
                       + "AppleWebKit/537.36 (KHTML, like Gecko) "
                       + "Chrome/79.0.3945.88 Safari/537.36 Edg/79.0.309.56")
    return [request]

def _jump_to_first_not_processed_row(sheet: openpyxl.worksheet.worksheet.Worksheet) -> int or None:
    '''Find the first row in the spreadsheet that is not processed (has 'F' in the 'Processed' row).'''
    for num in range(2, sheet.max_row + 1):
        if sheet[f"P{num}"].value == "F":
            return num
    return None

def _update_spreadsheet(row: int, info: dict, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    '''Change the information on the spreadsheet to the one listed on WebSOC.'''
    sheet["C" + str(row)].value = info["dept_name"]
    sheet["D" + str(row)].value = info["course_number"]
    sheet["E" + str(row)].value = info["course_code"]
    sheet["F" + str(row)].value = info["course_title"]
    sheet["G" + str(row)].value = "; ".join(info["instructors"])
    sheet["P" + str(row)].value = "T"

def _clean_data() -> None:
    '''Begin to send the request by one course code at a time and parse the data from WebSOC.'''
    logging.basicConfig(level=logging.DEBUG)
    for spreadsheet_file in SPREADSHEET_FILES:
        try:
            file = openpyxl.load_workbook("../temp/" + spreadsheet_file, data_only=True)
            with open("log.txt", "w") as error_log:
                for sheetname in file.sheetnames:
                    sheet = file[sheetname]
                    start = _jump_to_first_not_processed_row(sheet)
                    
                    if start is not None:
                        while start <= sheet.max_row:
                            course_code = str(sheet["E" + str(start)].value) if len(str(sheet["E" + str(start)].value)) == 5 else "0" + str(sheet["E" + str(start)].value)
                            logging.info(f"[{asctime()}] Processing course #{course_code} ({sheetname}).")

                            info = None
                            if sheetname.split()[0] != "Summer":
                                request = _build_non_summer_request({
                                    "quarter": sheetname.split()[1] + "-" + TERM_DICT[sheetname.split()[0].upper()],
                                    "course_code": course_code
                                })
                                info = _get_data(request, course_code)
                            else:
                                requests = _build_summer_requests({
                                    "year": sheetname.split()[1],
                                    "course_code": course_code
                                })
                                info = _get_data(requests, course_code)

                            if sheet["O" + str(start)].value is None:
                                sheet["O" + str(start)].value = 0

                            if info["success"] == True:
                                _update_spreadsheet(start, info, sheet)
                                file.save("../processed_data/_" + spreadsheet_file)
                            else:
                                logging.warning(f"[{asctime()}] Failed to process course #{course_code} ({sheetname}): " + info["error_message"])
                                error_log.write(f"[{asctime()}] Failed to process course #{course_code} ({sheetname}): {info['error_message']}\n")
                            start += 1

                            pause_time = uniform(5, 9)
                            logging.info(f"[{asctime()}] Halt the process for {pause_time:.2f} seconds to protect the WebSOC server.")
                            sleep(pause_time)
        except KeyboardInterrupt:
            logging.info(f"[{asctime()}] Exiting the program...")
        except Exception as exception:
            logging.warning(f"[{asctime()}] Error encountered: {exception}")

if __name__ == "__main__":
    _clean_data()
